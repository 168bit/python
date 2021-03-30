import os, win32com.client as win32, datetime, time, smtplib
from openpyxl import load_workbook
from email.message import EmailMessage
        
def fix_excel(name):
    try:

        if name=="H:\Thumbs.db" or "H:\~$" in name: return
        
        print("Check File " + name)

        workbook = load_workbook(filename=name, read_only=False)
        sheet = workbook.active
        rowcount = 0
        columncount = sheet.max_column
        r_num = 1

        for i in range(1, sheet.max_row+1):	
            if "袋模编号" in str(sheet.cell(row=i, column=1).value) or "袋模编号" in str(sheet.cell(row=i, column=2).value):
                rowcount = i
        
        #fix row number
        for i in range(8, rowcount):
            sheet.cell(row=i, column=1,value=r_num)
            r_num = r_num + 1 

        #fix header
        for i in range(4, columncount):		

            if "客户" in str(sheet.cell(row=4, column=i).value):
                if sheet.cell(row=4, column=i+1).value is not None:
                    sheet.cell(row=4, column=i, value = "客户：" + str(sheet.cell(row=4, column=i+1).value))
                    sheet.cell(row=4, column=i+1, value = '')
                    print("Fixed Buyer")   
            elif "款号" in str(sheet.cell(row=4, column=i).value):
                if sheet.cell(row=4, column=i+1).value is not None:
                    sheet.cell(row=4, column=i, value = "款号：" + str(sheet.cell(row=4, column=i+1).value))
                    sheet.cell(row=4, column=i+1, value = '')
                    print("Fixed ")
            elif "款式" in str(sheet.cell(row=4, column=i).value):
                if sheet.cell(row=4, column=i+1).value is not None:
                    sheet.cell(row=4, column=i, value = "款式：" + str(sheet.cell(row=4, column=i+1).value))
                    sheet.cell(row=4, column=i+1, value = '')
                    print("Fixed Style")
                
            elif "成衣" in str(sheet.cell(row=4, column=i).value):
                if sheet.cell(row=4, column=i+1).value is not None:
                    sheet.cell(row=4, column=i, value = "成衣：" + str(sheet.cell(row=4, column=i+1).value))
                    sheet.cell(row=4, column=i+1, value = None)
                    print("Fixed PONo")
        
        #fix footer    
        for i in range(1, columncount-4):
            if "#" in str(sheet.cell(row=rowcount, column=i).value):
                sheet.cell(row=rowcount, column=i, value = "")
                print("Fixed #")

        workbook.save(name)
    except:
        print("File is open, Skip")
        return

def count_files(dir):
    return len([1 for x in list(os.scandir(dir)) if x.is_file()])

def send_mail():
    msg = EmailMessage()
    msg['Subject'] = 'QA-MO Import'
    msg['From'] = 'ycpost_ipps@crystal-ydcl.com.kh'
    msg['To'] = 'loek_sarith@crystal-ydcl.com.kh'
    msg.set_content("Dear Reciepient, Please Check QA-MO Folder")
    try:
        smtpObj = smtplib.SMTP('192.168.106.22')
        smtpObj.send_message(msg) 
        print("Send Mail at " + str(datetime.datetime.now().time()))
    except SMTPException:
        print("Error: unable to send email")

basepath = 'C:\VDrive\Python\excel'
cf = 0
while True:
    
    if cf == count_files(basepath) : send_mail()

    if count_files(basepath) > 1 :
        for f_name in os.listdir(basepath):
            if os.path.isfile(os.path.join(basepath, f_name)):
                fix_excel(basepath + "\\" + f_name)
        cf = count_files(basepath)
    
    print("Sleep 1h")
    time.sleep(1800)
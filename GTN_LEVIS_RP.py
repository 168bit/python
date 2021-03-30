from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
import time
import os
import sys
from openpyxl import Workbook

def scrap_data(id,pwd,po):
    barcode = []
    options = Options()
    try:
        print("Loading Chrome...")
        options.binary_location="C:\\VDrive\\GTN_LEVIS\\config\\chrome-win\\chrome.exe"
        print("Loading Exentions...")
        options.add_extension("C:\\VDrive\\GTN_LEVIS\\config\\GTN_1_3_0_0.crx")
        options.add_argument("--headless")
        options.add_argument("window-size=1920x1080")
        options.add_argument("--disable-gpu-vsync")
        print("Loading Driver...")
        driver = webdriver.Chrome(executable_path="C:\\VDrive\\GTN_LEVIS\\config\\chromedriver.exe", options=options)
        print("Loading GTN...")
        driver.get("https://network.gtnexus.com/en/trade/login.jsp")
        driver.find_element_by_id("login").send_keys(id)
        driver.find_element_by_id("password").send_keys(pwd)
        driver.find_element_by_id("loginButton").click()
        driver.find_element_by_id("navmenu__applications").click()
        driver.find_element_by_id("navmenu__inprogressmanifestsprintscanship").click()
        
        time.sleep(2)
        print("Loading Print Scan Ship...")
        driver.find_element_by_link_text("Print Package Labels").click()
        driver.find_element_by_name("poNum").send_keys(po)
        driver.find_element_by_xpath("//div[label = 'Reprint Individual Packages']").click()
        driver.find_element_by_xpath("//*[button = 'OK']").click()
        time.sleep(3)
        print("Getting Reprint Label...")
        count = driver.find_elements_by_xpath("//div[@class='x-grid-group-body']/div")
        if len(count) < 1:
            print("Error: Wrong PO Number or Cannot Connect to TradCard")
            time.sleep(5)
            sys.exit(0)
        for i in range(len(count)):
            barcode.append(driver.find_element_by_xpath("//div[@class='x-grid-group-body']/div["+str(i+1)+"]/table/tbody/tr/td[6]/div").text)
        return barcode
    except:
        print("Error Execption")
        time.sleep(60)
def gen_excel(barcode,po,so):
    print("Generate Excel...")
    filepath = "C:\\Users\\"+os.getlogin()+"\\Desktop\\PO_"+str(po)+"_REPRINT.xlsx"
    wb = Workbook() 
    sheet = wb.active
    i = 1
    j = 1
    sheet.cell(row=1, column=1, value = "SO")
    sheet.cell(row=1, column=2, value = "PO")
    sheet.cell(row=1, column=3, value = "Label")
    for bar in barcode:
        x = bar.find("-")
        if x > 0:
            bar1 = bar[:x]
            bar2 = bar[x+1:]
            for j in range(int(bar2)-int(bar1)):
                i = i+1
                sheet.cell(row=i, column=1, value = so)
                sheet.cell(row=i, column=2, value = po)
                sheet.cell(row=i, column=3, value = int(bar1)+j)
                j = j+1
            i = i+1
            sheet.cell(row=i, column=1, value = so)
            sheet.cell(row=i, column=2, value = po)
            sheet.cell(row=i, column=3, value = bar2)
        else:
            i = i+1
            sheet.cell(row=i, column=1, value = so)
            sheet.cell(row=i, column=2, value = po)
            sheet.cell(row=i, column=3, value = bar)
    wb.save(filepath)
    print("=================================SUCCESS=================================")
    time.sleep(3)

def load_cre(file):
    print("Loading Config...")
    with open(file) as f:
        data = f.read().splitlines()
    user = data[0][5::]
    pwd = data[1][5::]
    return (user,pwd)

#start
print("=================================START=================================")
#load credential
cre = load_cre("C:\\VDrive\\GTN_LEVIS\\config\\config.txt")
usr = cre[0]
pwd = cre[1]
#get PO Input
so = input("Enter SONo: ")
po = input("Enter PONo: ")
#gen excel
gen_excel(scrap_data(usr,pwd,po),po,so)


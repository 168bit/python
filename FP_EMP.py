from zk import ZK, const
import pyodbc, datetime, time, smtplib, os,  win32com.client as win32
from openpyxl import load_workbook
from email.message import EmailMessage

conn_sql = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
                        'Server=192.168.109.4;'
                        'Database=YDCLFMISDB;'
                       "uid=fmisuser;pwd=ydclsqladmin")
conn_fp = None

class FP_List(object):
    d_id = ""
    d_name = ""
    d_ip = ""
    def __init__(self,d_id,d_name,d_ip):
        self.d_id = d_id
        self.d_name = d_name
        self.d_ip = d_ip

class FP_Data(object):
    id = ""
    emp_no = ""
    s_date = ""
    s_time = ""
    def __init__(self,id,emp_no,s_date,s_time):
        self.id = id
        self.emp_no = emp_no
        self.s_date = s_date
        self.s_time = s_time

def load_fp(filepath):
    try:
        workbook = load_workbook(filename=filepath, read_only=False)
        sheet = workbook.active
        FP_obj = []
        for i in range(2, sheet.max_row+1):
            d_id = str(sheet.cell(row=i, column=1).value)
            d_name = str(sheet.cell(row=i, column=2).value)
            d_ip = str(sheet.cell(row=i, column=3).value)
            FP_obj.append(FP_List(d_id,d_name,d_ip))
        return FP_obj
    except:
        print("Can't Find Fingerprint File")

def connect_fp(ip):
    Att_obj = []
    zk = ZK(ip, port=4370, timeout=5, password=636416, force_udp=False, ommit_ping=False)
    try:
        conn_fp = zk.connect()
        record = conn_fp.get_attendance()
        for i in range(0, len(record)):
            ID = str(record[i]).split()[1]
            Scan_Date = str(record[i]).split()[3] + " 00:00:00"
            Scan_Time = str(record[i]).split()[4]
            Scan_Time = Scan_Time.replace(':','')
            Att_obj.append(FP_Data("02",ID, Scan_Date, Scan_Time))
        return Att_obj

    except Exception as e:
        if "timed out" in str(e) :
            r = send_mail(ip)
            return r

def insert_sql(data):
    r = "SUCCESS"
    rc = 0
    sql = 'Insert into AttMachine_Test values (?,?,?,?,NULL,NULL,NULL)'
    for i in range(0, len(data)):
        try:
            cursor.execute(sql,data[i].id,data[i].emp_no,data[i].s_date,data[i].s_time)
            rc = rc + cursor.rowcount
        except Exception as e:
            r="FAILED"
    conn_sql.commit()
    #clear figerprint
    return r + "(" + str(rc) + ")"

def send_mail(ip):
    msg = EmailMessage()
    msg['Subject'] = 'Finger Print Import'
    msg['From'] = 'ycpost_nms@crystal-ydcl.com.kh'
    msg['To'] = 'loek_sarith@crystal-ydcl.com.kh'
    msg.set_content("Dear Reciepient, Please Check Finger Print IP: " + str(ip))
    try:
        smtpObj = smtplib.SMTP('192.168.106.22')
        smtpObj.send_message(msg)
        return -1
    except Exception as e:
        return -2

def main_task():
    cwd = os.getcwd()
    Device_List = load_fp(cwd+"\\config\\FP_List.xlsx")
    print("=========================================START==================================================")
    print("Device_Name" + "\t" + "IP_Address" + "\t" + "Log_Count" + "\t" + "Log_Time" + "\t\t" + "Insert FMIS")
    for i in range(0, len(Device_List)):
        data = connect_fp(Device_List[i].d_ip)
        if data == None:
            print(Device_List[i].d_id + "-" + Device_List[i].d_name + "\t\t" + Device_List[i].d_ip + "\t" + "No Log" + "\t\t" + str(datetime.datetime.now())[:19] + "\t" + "NONE")
        elif data == -1:
            print(Device_List[i].d_id + "-" + Device_List[i].d_name + "\t\t" + Device_List[i].d_ip + "\t" + "Send_M" + "\t\t" + str(datetime.datetime.now())[:19] + "\t" + "NONE")
        elif data == -2:
            print(Device_List[i].d_id + "-" + Device_List[i].d_name + "\t\t" + Device_List[i].d_ip + "\t" + "Fail_M" + "\t\t" + str(datetime.datetime.now())[:19] + "\t" + "NONE")
        else:
            fp_count=len(data)
            status = insert_sql(data)
            print(Device_List[i].d_name + "\t\t" + Device_List[i].d_ip + "\t" + str(fp_count) + "\t\t" + str(datetime.datetime.now())[:19] + "\t" + str(status))
    print("=======================================END==============================================")    

while True:
    cursor = conn_sql.cursor()
    main_task()
    conn_sql.close()
    time.sleep(3600)
